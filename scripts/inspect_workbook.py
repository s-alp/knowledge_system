from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def to_jsonable(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("usage: inspect_workbook.py <xlsx_path>")

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, data_only=False)

    result: dict[str, object] = {
        "path": str(workbook_path),
        "sheet_names": wb.sheetnames,
        "active_sheet": wb.active.title,
        "sheets": {},
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[object | None]] = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            rows.append([to_jsonable(value) for value in row])
        result["sheets"][sheet_name] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_rows": rows,
        }

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
