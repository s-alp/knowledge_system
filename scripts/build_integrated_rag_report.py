from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
SOURCE_PATH = WORKSPACE / "output" / "spreadsheet" / "rag_test_summary_unified_2026-04-23.xlsx"
TARGET_PATH = WORKSPACE / "output" / "spreadsheet" / "rag_test_summary_unified_2026-04-23.xlsx"

INTEGRATED_ROWS = [
    [
        "I01",
        "根拠資料の混線",
        "T02, T04, T06, T10, T11, T14, T15, T18",
        "別案件・別カテゴリ・負例時の無関係資料が根拠に混ざる。追撃で条件を足しても、条件外資料が根拠集合に残る。",
        "案件名・客先名・カテゴリ名・除外条件を満たさない資料を再ランキング前に除外し、回答本文と参考資料の整合チェックを追加する。",
        "条件不一致の資料を根拠候補から外す仕組みはありますか。負例時に『参考なし』を返す設計は可能ですか。",
        "高",
        "質問で指定した案件・カテゴリに直接対応する資料群。T18 であれば `20241202_対応項目.txt` のような項目列挙文書が本来の着地点。",
    ],
    [
        "I02",
        "追撃条件の再反映不足",
        "T13, T14, T15",
        "用語定義、除外条件、出力形式指定を追加しても、初回の検索・根拠集合を十分に捨て切れず、追撃後の条件が弱く反映される。",
        "追撃ごとに再検索・再絞り込みを行い、直前の回答ではなく最新条件を優先する。追撃後の根拠集合を再計算する。",
        "追撃で追加された条件を、初回の検索結果ではなく最新条件として強制反映できますか。",
        "高",
        "追撃で追加した条件を最も素直に満たす資料。除外条件つきなら、その条件を満たす同案件資料だけに絞った根拠集合。",
    ],
    [
        "I03",
        "案件名・客先名・カテゴリ名の切り分け不足",
        "T02, T11, T14, T18",
        "『広島アルミ』『ガントリー』『改造対応項目』のような案件・カテゴリ条件を入れても、同一案件内の別カテゴリや別案件へ流れる。",
        "客先名・案件名・装置カテゴリ・文書種別を独立メタデータとして持ち、質問条件と一致しない資料のスコアを大きく下げる。",
        "案件内・企業内の資料でも、カテゴリ不一致資料を根拠から除外する設定は可能ですか。",
        "高",
        "質問で指定した案件・客先・カテゴリに一致する資料。広島アルミなら広島アルミ配下、ガントリーならガントリー系資料。",
    ],
    [
        "I04",
        "BOM・部品情報から案件名へ上がれない",
        "T08, T09, T12, T15",
        "部品名や用途名までは拾えても、案件名・客先名へ正規化して返せない。案件列挙をしても根拠が各案件に対応しない。",
        "部品名→構成→案件名の連結を強化し、案件列挙時は案件ごとに最低1件の根拠を必須にする。根拠がない案件は除外する。",
        "部品起点の検索で案件名へ上げられない場合、案件未特定として返す制御や、案件ごとの根拠必須化は可能ですか。",
        "高",
        "案件名を直接裏付けられる BOM / 構成表 / チェックリスト。列挙した案件ごとに 1 件ずつ必要。",
    ],
    [
        "I05",
        "規格表・列挙項目の抽出精度",
        "T17",
        "正式資料には到達するが、表の列挙項目を言い換えたり取り違えたりする。今回も『熱処理・硬度・硬化深さ』が正しく抜けていない。",
        "表形式の規格から列挙を求められた場合は、表見出し・列項目を優先抽出し、要約前に原文一致チェックを入れる。",
        "表形式PDFや表形式テキストで、列挙対象の見出し・項目を優先抽出する設定は可能ですか。",
        "中",
        "該当規格の表や見出し。T17 では `熱処理指定方法_2GDE82D.pdf` の表1 が本来の着地点。",
    ],
    [
        "I06",
        "Excel・Word仕様書の読取り精度",
        "T19, T20, T21",
        "Excel/Word 仕様書への質問では、値が合う場合もあるが、表セルや手順書本文そのものに着地できず、別資料根拠や逆転回答が出る。特に Word 手順書の数値回答は崩れやすい。",
        "Excel は対象シート・対象行列、Word は見出し配下の段落本文を優先抽出する。文書種別を条件に使い、別種別資料への誤着地を抑える。",
        "Excel/Word 仕様書に対して、文書種別と表・段落位置を優先する抽出制御は可能ですか。",
        "高",
        "対象の Excel シートや Word 手順書本文そのもの。T19 は `STEP2_＞STEP3ハンドの仕様r2.xlsx`、T20 は `C800-02_ミスト・集塵装置クミ品質要求事項確認チェックリスト.xlsx`、T21 は `maintenance_（gantry）-スマートライン2_最新.doc`。",
    ],
]

SUCCESS_ROWS = [
    [
        "S01",
        "T03",
        "コマツ小山のガントリーの走行速度",
        "案件名とカテゴリ名が明確で、資料パスにも案件名・ガントリーが含まれると安定して正答した。",
    ],
    [
        "S02",
        "T07",
        "正式資料名の直接指定",
        "正式ファイル名を直接指定すると、資料の存在確認とページ特定は安定している。",
    ],
    [
        "S03",
        "T16",
        "澁谷工業の長穴寸法",
        "正式規格名と対象サイズが明確な質問では、数値と根拠ページがローカル資料と整合した。",
    ],
]


def apply_header_style(ws, row_idx: int, fill_color: str = "1F4E78") -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_body_style(ws, start_row: int = 2) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def reset_sheet(ws) -> None:
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)


def build_integrated_sheet(ws) -> None:
    headers = [
        "統合ID",
        "統合論点",
        "対象テストID",
        "統合した事実",
        "改善提案",
        "開発先確認事項",
        "優先度",
        "本来の想定根拠",
    ]
    ws.append(headers)
    for row in INTEGRATED_ROWS:
        ws.append(row)
    apply_header_style(ws, 1)
    apply_body_style(ws)
    set_widths(
        ws,
        {
            "A": 10,
            "B": 28,
            "C": 24,
            "D": 56,
            "E": 52,
            "F": 48,
            "G": 10,
            "H": 56,
        },
    )


def build_success_sheet(ws) -> None:
    headers = ["成功ID", "参照テストID", "質問テーマ", "成功パターン"]
    ws.append(headers)
    for row in SUCCESS_ROWS:
        ws.append(row)
    apply_header_style(ws, 1, fill_color="548235")
    apply_body_style(ws)
    set_widths(ws, {"A": 10, "B": 14, "C": 32, "D": 64})


def main() -> int:
    wb = load_workbook(SOURCE_PATH)

    if "統合論点" in wb.sheetnames:
        reset_sheet(wb["統合論点"])
        ws_integrated = wb["統合論点"]
    else:
        ws_integrated = wb.create_sheet("統合論点", 1)

    if "成功例" in wb.sheetnames:
        reset_sheet(wb["成功例"])
        ws_success = wb["成功例"]
    else:
        ws_success = wb.create_sheet("成功例", 2)

    build_integrated_sheet(ws_integrated)
    build_success_sheet(ws_success)

    TARGET_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET_PATH)
    print(TARGET_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
