from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
SOURCE_PATH = WORKSPACE / "output" / "spreadsheet" / "rag_test_summary_reconciled_2026-04-23.xlsx"
TARGET_PATH = WORKSPACE / "output" / "spreadsheet" / "rag_test_summary_unified_2026-04-23.xlsx"

EXPECTED_SOURCES = {
    "T01": (
        r"shibuya_specs\熱処理指定方法_2GDE82D.pdf / shibuya_specs\材料の熱処理適性および硬さ_2GDE81D.pdf",
        "『SESの熱処理規格』は本来、澁谷工業の熱処理系規格へ寄るべきで、ローカル検証素材では 2GDE82D と 2GDE81D が直接候補になる。",
    ),
    "T02": (
        r"shibuya_specs\長穴の寸法_2GDE03D.pdf",
        "M5 に対するボルト穴サイズは、ローカル検証素材では『長穴の寸法_2GDE03D.pdf』の表が最も直接の根拠になる。",
    ),
    "T03": (
        r"komatsu_koyama_gantry\【製作仕様書】20220404_KO小山ガントリー装置STEP2.pdf",
        "コマツ小山ガントリーの走行速度は、ローカルの製作仕様書 PDF から出るはずの情報で、案件名と装置名も一致している。",
    ),
    "T04": (
        "なし",
        "『トヨタ向けのガントリー資料』は今回の対象やローカル検証素材に存在しないため、本来は『資料なし』で返るべき。",
    ),
    "T05": (
        "ローカル検証素材なし（広島アルミの治具資料は未回収）",
        "この質問に答えるには広島アルミ配下の治具資料が必要だが、今回のローカル検証素材には治具資料を持ち込んでいない。",
    ),
    "T06": (
        r"shibuya_specs\熱処理指定方法_2GDE82D.pdf / shibuya_specs\材料の熱処理適性および硬さ_2GDE81D.pdf",
        "『澁谷工業のSES規格』として資料名を列挙するなら、少なくともローカルにある熱処理系規格 2 件へ寄るべき。",
    ),
    "T07": (
        r"shibuya_specs\熱処理指定方法_2GDE82D.pdf",
        "質問文で正式ファイル名を直接指定しており、ローカルにも同名 PDF があるため、この資料に着地するのが自然。",
    ),
    "T08": (
        "ローカル検証素材なし（案件別 BOM / 構成表が必要）",
        "SMC のシリンダー使用案件を答えるには、案件ごとの BOM や構成表が必要だが、今回のローカル検証素材には十分な案件横断 BOM を持ち込んでいない。",
    ),
    "T09": (
        "ローカル検証素材なし（案件別 BOM / 構成表が必要）",
        "案件名だけを列挙するには、部品名から案件名へ上がれる案件別 BOM / 構成表が必要で、単一チェックリストでは不足する。",
    ),
    "T10": (
        "なし",
        "『澁谷工業のガントリー資料』は今回のローカル検証素材にも対象範囲にも見当たらないため、本来は『確認できない』が正しい。",
    ),
    "T11": (
        r"hiroshima_alumi_gantry\現地改造計画図.pdf / hiroshima_alumi_gantry\ケースライン確認図.pdf",
        "広島アルミのガントリー関連資料名を問う質問なので、ローカルの広島アルミガントリー関連 PDF に着地するのが自然。",
    ),
    "T12": (
        r"T:\NTC\TF設計部\TD2\コマツ小山\01_治具\第二期治具\140ワーク計画\0113_140ワーク_KOSMEKホールクランプ当たり資料.xlsx / T:\NTC\TF設計部\TD2\コマツ小山\01_治具\第二期治具\140ワーク計画\0117_140ワーク_KOSMEKホールクランプ当たり資料.xlsx",
        "KOSMEK 油圧シリンダの案件を問うなら、社内パス上で確認できた KOSMEK ホールクランプ当たり資料のような案件別資料に着地するのが妥当。",
    ),
    "T13": (
        r"shibuya_specs\熱処理指定方法_2GDE82D.pdf",
        "追撃で『澁谷工業の熱処理規格だけ』と明示しているため、ローカルでは 2GDE82D が本命の想定根拠になる。",
    ),
    "T14": (
        r"hiroshima_alumi_gantry\現地改造計画図.pdf / hiroshima_alumi_gantry\ケースライン確認図.pdf",
        "質問条件が『広島アルミ』『ガントリー／ハンド』『側部カバー除外』なので、コマツ小山資料へ飛ばず、広島アルミ側のガントリー系資料に着地するのが自然。",
    ),
    "T15": (
        "案件名を裏付けられる BOM / 構成表 / チェックリスト。少なくとも列挙した案件ごとに 1 件ずつ対応根拠が必要。",
        "この質問は『案件名の列挙』が目的なので、案件を返すなら各案件を直接裏付ける資料が必要。広島アルミ側部カバーのチェックリスト 2 件だけでは不十分。",
    ),
    "T16": (
        r"shibuya_specs\長穴の寸法_2GDE03D.pdf",
        "質問文に M5・長穴・板厚が明示されており、ローカル規格 PDF に直接該当表がある。",
    ),
    "T17": (
        r"shibuya_specs\熱処理指定方法_2GDE82D.pdf",
        "質問文に正式規格名相当の『熱処理指定方法』が含まれ、ローカル PDF の表1 に直接『指定すべき項目』が載っている。",
    ),
    "T18": (
        r"hiroshima_alumi_gantry\20241202_対応項目.txt",
        "質問文そのものが『改造対応項目』で、ローカル txt に対応項目が箇条書きで直接列挙されている。",
    ),
    "T19": (
        r"excel_word_specs\komatsu_koyama\STEP2_＞STEP3ハンドの仕様r2.xlsx",
        "質問文に『ガントリーハンド仕様』『車種107』『前半』『STEP3移行時』が含まれ、ローカル Excel の表セルに直接対応項目がある。",
    ),
    "T20": (
        r"excel_word_specs\hiroshima_alumi\C800-02_ミスト・集塵装置クミ品質要求事項確認チェックリスト.xlsx",
        "PJ No.・形格・組図名称がローカル Excel の冒頭行にまとまっており、この資料だけで回答可能。",
    ),
    "T21": (
        r"excel_word_specs\komatsu_koyama\maintenance_（gantry）-スマートライン2_最新.doc",
        "質問文が『保守要領』『X軸走行用サーボモーター交換』『バックラッシュ狙い値』なので、ローカル Word 手順書本文の該当段落から出るべき内容。",
    ),
}

NEW_ROWS = [
    [
        "T14",
        "追撃改善",
        "広島アルミのガントリー関連資料として、側部カバーや干渉チェックは除外してください。ガントリーまたはハンド関係の資料名だけを列挙してください。",
        "http://210.165.3.139/web/chat/e143eea3-2152-434d-9af5-6e03abca5fd7",
        "除外条件を追加しても、コマツ小山の『20220711_ガントリーハンド資料/新しいテキスト ドキュメント.txt』を返し、根拠もコマツ小山ガントリー資料へ混線した。",
        "『広島アルミ』『ガントリー』『側部カバー除外』を明示しても、客先境界とカテゴリ境界の両方を守れなかった。除外指定後の再検索でも別案件資料が根拠に混入している。",
        "追撃で追加した除外条件と客先条件を、回答生成前の根拠集合に強制反映する。条件に合わない案件やカテゴリの資料は再ランキング前に落とす。",
        "追撃時に『除外条件』と『客先条件』を初回検索結果から再評価し、別案件資料を根拠から除外する仕組みはありますか。",
        "高",
        "問題あり",
        *EXPECTED_SOURCES["T14"],
    ],
    [
        "T15",
        "追撃改善",
        "案件名を答えてください。部品名や用途名ではなく、客先名や案件名を挙げてください。SMCのシリンダーが使われている案件だけを列挙してください。",
        "http://210.165.3.139/web/chat/e1d985e6-c078-41b5-83ae-600020158621",
        "出力形式指定で『案件名』の列挙には寄ったが、広島アルミ・コマツ小山・No.92を挙げつつ、根拠は広島アルミ側部カバーのチェックリスト2件のみだった。",
        "追撃で出力形式を指定すると案件名らしき回答は返るが、列挙した案件ごとの根拠整合は崩れたままだった。コマツ小山やNo.92を裏付ける資料が提示されていない。",
        "列挙系の出力形式指定時は、各案件ごとに最低1件の根拠を必須にする。根拠が取れない案件名は回答から除外する。",
        "『案件名だけを列挙』のような出力形式指定時に、各項目へ対応する根拠を必須化できますか。",
        "高",
        "問題あり",
        *EXPECTED_SOURCES["T15"],
    ],
    [
        "T16",
        "規格検索",
        "澁谷工業の規格で、M5に対応する長穴の幅Aと最大板厚Tを教えてください。資料名とページも示してください。",
        "http://210.165.3.139/web/chat/2c09bad6-5895-42d2-9270-d3dbc6dad583",
        "M5の長穴幅A=6 mm、最大板厚T=12 mmと回答し、『長穴の寸法_2GDE03D.pdf』ページ1を根拠表示した。",
        "ローカル資料の『長穴の寸法_2GDE03D.pdf』と整合した。正式な規格名と数値指定が明確なため、規格検索として安定して回答できている。",
        "寸法規格のように正式資料名と対象サイズが明確な質問パターンを、設計実務向けの成功例として整理する。",
        "",
        "低",
        "良好",
        *EXPECTED_SOURCES["T16"],
    ],
    [
        "T17",
        "規格検索",
        "澁谷工業の熱処理指定方法で、部品図に指定すべき項目を列挙してください。資料名とページも示してください。",
        "http://210.165.3.139/web/chat/568f633a-3261-48f5-945f-e6299d40498d",
        "『熱処理箇所の明示』『指定名称の記載』を回答し、『熱処理指定方法_2GDE82D.pdf』ページ1を根拠表示した。",
        "ローカル資料の正式規格名には到達したが、原文の『指定すべき項目』は熱処理・硬度・硬化深さであり、回答では備考相当の『熱処理箇所の明示』へ置き換わっていた。部分正答だが項目抽出は不正確だった。",
        "正式資料名に到達した後の回答生成で、表の列挙項目をそのまま抜き出す制御を強める。要約で項目を言い換える前に、列挙対象の原文一致チェックを入れる。",
        "表形式の規格から『列挙』を求めたときに、表の見出し・列項目を優先して抽出する設定はできますか。",
        "中",
        "要改善",
        *EXPECTED_SOURCES["T17"],
    ],
    [
        "T18",
        "案件検索",
        "広島アルミの改造対応項目として挙がっているものを列挙してください。根拠資料名も示してください。",
        "http://210.165.3.139/web/chat/fb103f1e-cd92-4132-9404-a8344bb75323",
        "『戻しボタン』『切削条件』『製図作業』を回答し、根拠に『C300-09_ツーリング品質要求事項確認チェックリスト.xlsx』を提示した。",
        "ローカルの『20241202_対応項目.txt』にはワークハンド、ハンド置台、踏台、安全柵、ロボット制御盤架台、トイ、クーラント2次タンク移設用トイ・配管が列挙されており、回答内容は整合しない。広島アルミ案件内でも別資料へ誤着地している。",
        "案件内の複数資料から拾う場合でも、質問文に近いファイル名や受領メモを優先する。『対応項目』『列挙』系ではチェックリストより先に箇条書き文書を優先する再ランキングが有効。",
        "案件内に複数資料がある場合、質問語と一致するファイル名やテキストメモを優先する重み付けは可能ですか。",
        "高",
        "問題あり",
        *EXPECTED_SOURCES["T18"],
    ],
    [
        "T19",
        "仕様書検索",
        "コマツ小山のガントリーハンド仕様で、車種107の前半はSTEP3移行時にツメの新製が必要ですか。根拠資料名も示してください。",
        "http://210.165.3.139/web/chat/45cc17d9-c03f-442b-8185-168423b38940",
        "『不要』と回答したが、根拠は別ExcelやPDFに流れ、ローカルの『STEP2_＞STEP3ハンドの仕様r2.xlsx』とは整合しなかった。",
        "ローカルExcelでは車種107前半の『STEP3移行時 ツメの新製』は『必要』となっており、回答本文が逆転している。ハンド仕様表そのものに着地していない可能性が高い。",
        "Excel仕様表を対象にした質問では、該当シートの行・列を特定してから回答生成する。別資料を横断参照する前に、質問文に近い仕様表を優先する。",
        "Excel の表形式仕様に対して、対象シート・対象行列を優先抽出する制御は可能ですか。",
        "高",
        "問題あり",
        *EXPECTED_SOURCES["T19"],
    ],
    [
        "T20",
        "仕様書検索",
        "広島アルミのミスト・集塵装置クミのPJ No.、形格、組図名称を教えてください。根拠資料名も示してください。",
        "http://210.165.3.139/web/chat/11de51e7-c775-46b2-a834-ba98b918821c",
        "PJ No.=00K1A4980、形格=FTL-2184GC、組図名称=ミスト・集塵装置クミと回答。値自体はローカルExcelと整合したが、参考に別案件や別資料が大量混入した。",
        "ローカルの『C800-02_ミスト・集塵装置クミ品質要求事項確認チェックリスト.xlsx』と数値・名称は一致した。一方で、根拠としてマガジン改造やツーリング資料まで混ざり、根拠集合は過剰だった。",
        "値が正しい場合でも、根拠資料は質問に直接対応する仕様書へ絞る。代表根拠1〜2件に制限し、別案件資料は参考から落とす。",
        "正答時でも、代表根拠を質問に最も近い仕様書だけへ絞る設定はできますか。",
        "中",
        "要改善",
        *EXPECTED_SOURCES["T20"],
    ],
    [
        "T21",
        "仕様書検索",
        "コマツ小山ガントリーの保守要領で、X軸走行用サーボモーター交換時のバックラッシュ狙い値はいくつですか。根拠資料名も示してください。",
        "http://210.165.3.139/web/chat/6357f591-8477-4b47-8b20-bf25f7c77893",
        "バックラッシュ狙い値を『0.2』と回答したが、ローカルWordの保守要領では『0.45mm狙い』であり、参考資料も無関係だった。",
        "ローカルの『maintenance_（gantry）-スマートライン2_最新.doc』には『バックラッシュ0.45mm狙い』と明記されている。Word 手順書本文の数値抽出に失敗し、別Excelへ誤着地している。",
        "Word 手順書の見出しと段落本文を優先抽出し、数値回答では本文近傍の単位つき記載を優先する。Word と Excel を混在検索するときは文書種別を条件に使う。",
        "Word 手順書を含む検索で、本文近傍の数値と単位を優先し、別文書種別の資料を抑制する設定は可能ですか。",
        "高",
        "問題あり",
        *EXPECTED_SOURCES["T21"],
    ],
]

PENDING_UPDATES = {
    "広島アルミのガントリー関連資料として、側部カバーや干渉チェックを除外する追撃": (
        "回収済み",
        "追撃後もコマツ小山資料へ混線。T14 に整理。",
    ),
    "SMC案件に対する追加の出力形式指定追撃": (
        "回収済み",
        "案件名列挙は改善したが根拠整合は崩れたまま。T15 に整理。",
    ),
}


def clone_row_style(ws, source_row_idx: int, target_row_idx: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row_idx, column)
        target = ws.cell(target_row_idx, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)


def append_test_rows(ws) -> None:
    template_row = ws.max_row
    for row in NEW_ROWS:
        target_row = ws.max_row + 1
        ws.append(row)
        clone_row_style(ws, template_row, target_row)


def ensure_expected_source_columns(ws) -> None:
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if "期待資料" in headers:
        return
    ws.insert_cols(11, amount=2)
    ws.cell(1, 11).value = "期待資料"
    ws.cell(1, 12).value = "想定理由"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_side = Side(style="thin", color="D9E2F3")
    header_border = Border(left=header_side, right=header_side, top=header_side, bottom=header_side)
    for col in (11, 12):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = copy(ws.cell(1, 1).font)
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions["K"].width = 44
    ws.column_dimensions["L"].width = 48


def update_pending_sheet(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    resolved_fill = PatternFill("solid", fgColor="E2F0D9")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        content = row[1].value
        if content not in PENDING_UPDATES:
            continue
        status, note = PENDING_UPDATES[content]
        row[2].value = status
        row[3].value = note
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = resolved_fill


def populate_expected_columns(ws) -> None:
    for row_idx in range(2, ws.max_row + 1):
        test_id = ws.cell(row_idx, 1).value
        if test_id not in EXPECTED_SOURCES:
            continue
        expected_source, rationale = EXPECTED_SOURCES[test_id]
        ws.cell(row_idx, 11).value = expected_source
        ws.cell(row_idx, 12).value = rationale
        ws.cell(row_idx, 11).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row_idx, 12).alignment = Alignment(vertical="top", wrap_text=True)


def main() -> int:
    wb = load_workbook(SOURCE_PATH)
    ws_tests = wb["検証結果"]
    ws_pending = wb["未回収整理"]

    ensure_expected_source_columns(ws_tests)
    append_test_rows(ws_tests)
    populate_expected_columns(ws_tests)
    update_pending_sheet(ws_pending)

    TARGET_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET_PATH)
    print(TARGET_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
