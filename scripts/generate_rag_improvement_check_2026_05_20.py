from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
OUTPUT_PATH = WORKSPACE / "output" / "spreadsheet" / "rag_improvement_check_2026-05-20.xlsx"


ROWS = [
    {
        "id": "T17",
        "category": "規格表抽出",
        "question": "澁谷工業の熱処理指定方法で、部品図に指定すべき項目を列挙してください。資料名とページも示してください。",
        "url": "http://210.165.3.139/web/chat/63ae8ed4-c809-4d05-a0c7-69112f25e402",
        "answer": "すみません、わかりません。",
        "references": "熱処理指定方法_2GDE82D.pdf (ページ：1)",
        "expected": "熱処理指定方法_2GDE82D.pdf の表1。列挙対象は熱処理・硬度・硬化深さ。",
        "finding": "正しい資料に当たっているが、表項目を抽出できず回答不能。",
        "status": "後退",
        "priority": "高",
    },
    {
        "id": "T18",
        "category": "改造対応項目",
        "question": "広島アルミの改造対応項目として挙がっているものを列挙してください。根拠資料名も示してください。",
        "url": "http://210.165.3.139/web/chat/5da6629a-53e6-479b-8702-1a17dd90e20a",
        "answer": "すみません、わかりません。",
        "references": "C300-09_ツーリング品質要求事項確認チェックリスト.xlsx / 140ワーク下面仕上げ治具_C450-17_治具品質要求事項確認チェックリスト.xlsx",
        "expected": "hiroshima_alumi_gantry\\20241202_対応項目.txt",
        "finding": "本来の対応項目メモではなく、別案件・別カテゴリのチェックリストへ誤着地。",
        "status": "未改善",
        "priority": "高",
    },
    {
        "id": "T19",
        "category": "Excel仕様表",
        "question": "コマツ小山のガントリーハンド仕様で、車種107の前半はSTEP3移行時にツメの新製が必要ですか。根拠資料名も示してください。",
        "url": "http://210.165.3.139/web/chat/e2433f66-9df3-4647-acce-523c59fa92bf",
        "answer": "すみません、わかりません。",
        "references": "Step1_3比較_O2案_210803.pdf (ページ：4)",
        "expected": "excel_word_specs\\komatsu_koyama\\STEP2_＞STEP3ハンドの仕様r2.xlsx。答えは必要。",
        "finding": "誤答から回答不能にはなったが、対象Excel仕様表には着地していない。",
        "status": "要改善",
        "priority": "高",
    },
    {
        "id": "T20",
        "category": "Excel冒頭項目",
        "question": "広島アルミのミスト・集塵装置クミのPJ No.、形格、組図名称を教えてください。根拠資料名も示してください。",
        "url": "http://210.165.3.139/web/chat/c691b0f8-5775-4c11-9f8e-8bc6ebff751b",
        "answer": "すみません、わかりません。",
        "references": "C800-02_ミスト・集塵装置クミ品質要求事項確認チェックリスト.xlsx (シート名：CHECKLIST)",
        "expected": "PJ No.=00K1A4980、形格=FTL-2184GC、組図名称=ミスト・集塵装置クミ。",
        "finding": "正しい資料に当たっているが、4月時点で正答できた値を返せなくなった。",
        "status": "後退",
        "priority": "高",
    },
    {
        "id": "T21",
        "category": "Word手順書",
        "question": "コマツ小山ガントリーの保守要領で、X軸走行用サーボモーター交換時のバックラッシュ狙い値はいくつですか。根拠資料名も示してください。",
        "url": "http://210.165.3.139/web/chat/1f68b456-1a47-4f46-92a6-b38a08090674",
        "answer": "すみません、わかりません。",
        "references": "210827_スマートライン２ガントリ検討会.pdf (ページ：4)",
        "expected": "maintenance_（gantry）-スマートライン2_最新.doc。値は0.45mm狙い。",
        "finding": "Word手順書本文ではなく、別PDFへ誤着地。",
        "status": "要改善",
        "priority": "高",
    },
    {
        "id": "T16",
        "category": "寸法規格",
        "question": "澁谷工業の規格で、M5に対応する長穴の幅Aと最大板厚Tを教えてください。資料名とページも示してください。",
        "url": "http://210.165.3.139/web/chat/70056d20-ed92-488a-a498-7de0f66b9290",
        "answer": "すみません、わかりません。",
        "references": "長穴の寸法_2GDE03D.pdf (ページ：1)",
        "expected": "M5の長穴幅A=6 mm、最大板厚T=12 mm。",
        "finding": "4月時点で正答できた規格数値が、今回は正しい資料に当たっても回答不能。",
        "status": "後退",
        "priority": "高",
    },
    {
        "id": "T14",
        "category": "除外条件",
        "question": "広島アルミのガントリー関連資料として、側部カバーや干渉チェックは除外してください。ガントリーまたはハンド関係の資料名だけを列挙してください。",
        "url": "http://210.165.3.139/web/chat/f2f007d6-f8ab-42f6-9c18-fcb91eef9139",
        "answer": "すみません、わかりません。",
        "references": "20250227_広島アルミ_ドレン配管クミ_計画資料.xlsx / 20250117_広島アルミ_ミストコレクタクミ_計画資料.xlsx",
        "expected": "hiroshima_alumi_gantry\\現地改造計画図.pdf / hiroshima_alumi_gantry\\ケースライン確認図.pdf",
        "finding": "別客先混線は弱まったが、ガントリー/ハンド関連資料には着地しない。",
        "status": "一部改善・未解決",
        "priority": "中",
    },
    {
        "id": "T04",
        "category": "資料なし",
        "question": "トヨタ向けのガントリー資料はありますか。資料名を示してください。",
        "url": "http://210.165.3.139/web/chat/43c5e6d8-0d9c-43db-8f59-b389b6d9cc29",
        "answer": "すみません、わかりません。",
        "references": "140ワーク下面仕上げ治具_C450-17_治具品質要求事項確認チェックリスト.xlsx",
        "expected": "資料なし。参考資料も出さないことが望ましい。",
        "finding": "本文は資料なし寄りだが、無関係なコマツ治具資料を参考表示している。",
        "status": "一部改善・未解決",
        "priority": "中",
    },
    {
        "id": "SMC-1",
        "category": "BOM/部品起点",
        "question": "SMCのシリンダーが使われている案件を教えて。",
        "url": "http://210.165.3.139/web/chat/f13869f2-f050-4331-bea6-dfd50b4c3681",
        "answer": "コマツ小山、広島アルミ、コマツを列挙。",
        "references": "KO小山ガントリシリンダー・潤滑一覧.xlsx / X57-AK520C0-4E_TRANSFER DEV._ROBOT-BOM.xls / X57-AK24AC0-0L_PANEL A..xls",
        "expected": "案件名とSMC製シリンダーの対応を、案件ごとの根拠で示す。",
        "finding": "初回は候補を返すが、BOM中のSMC部品をシリンダーと断定しており厳密性が弱い。",
        "status": "要改善",
        "priority": "中",
    },
    {
        "id": "SMC-2",
        "category": "追撃・用語定義",
        "question": "ここでいうSMCはメーカー名です。部品名ではなく、SMC製シリンダーを使っている客先名や案件名を答えてください。",
        "url": "http://210.165.3.139/web/chat/f13869f2-f050-4331-bea6-dfd50b4c3681",
        "answer": "すみません、わかりません。",
        "references": "210827_社内アーム昇降構想打合せ.pdf",
        "expected": "初回根拠を再評価し、案件名と根拠の対応を改善する。",
        "finding": "追撃で初回より悪化し、無関係寄りのPDFへ飛ぶ。",
        "status": "未改善",
        "priority": "高",
    },
    {
        "id": "SMC-3",
        "category": "追撃・除外条件",
        "question": "用途名や部品名だけの回答は除外してください。根拠資料に案件名が確認できるものだけを列挙してください。",
        "url": "http://210.165.3.139/web/chat/f13869f2-f050-4331-bea6-dfd50b4c3681",
        "answer": "すみません、わかりません。",
        "references": "E2EZ-X4B1_2M_E2EZ-X8B1_2M...pdf / VQ7-6-OMD0001.pdf / 表面処理の種類・特性_2GXF22D.pdf",
        "expected": "用途名・部品名だけを落とし、根拠付き案件だけに絞る。",
        "finding": "条件追加後に根拠がさらに拡散。",
        "status": "未改善",
        "priority": "高",
    },
    {
        "id": "SMC-4",
        "category": "追撃・出力形式指定",
        "question": "表形式で、案件名、対象部品またはユニット名、根拠資料名、不明点を列にしてください。",
        "url": "http://210.165.3.139/web/chat/f13869f2-f050-4331-bea6-dfd50b4c3681",
        "answer": "広島アルミ、コマツ小山、広島アルミ/MEXを表で列挙。",
        "references": "干渉チェックエビデンスシート_ソクブカバークミ.xlsx / 干渉チェックエビデンスシート_OP140FIXTURE.xlsx / ボルト_2GDG30D.pdf",
        "expected": "SMCシリンダー条件を維持した表形式回答。",
        "finding": "表形式には従うが、SMC条件から外れた資料へ着地。",
        "status": "未改善",
        "priority": "高",
    },
]


SUMMARY_ROWS = [
    ("確認日時", "2026-05-20 12:01:24"),
    ("対象", "http://210.165.3.139/web/chat"),
    ("総評", "誤答抑制は強まった可能性があるが、正しい資料から本文を抽出できず、実務回答力は後退気味。"),
    ("主要問題1", "正しい資料に当たっても `すみません、わかりません。` になる。"),
    ("主要問題2", "`わかりません` の場合でも無関係資料が参考資料に出る。"),
    ("主要問題3", "追撃条件や出力形式指定で根拠資料が悪化する。"),
]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"


def build_workbook() -> Workbook:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "要約"
    ws_summary.append(["項目", "内容"])
    for row in SUMMARY_ROWS:
        ws_summary.append(row)
    style_sheet(ws_summary)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 90

    ws = wb.create_sheet("改善確認結果")
    headers = [
        "ID",
        "観点",
        "質問",
        "チャットURL",
        "今回の回答",
        "今回の参考資料",
        "期待結果",
        "事実",
        "判定",
        "優先度",
    ]
    ws.append(headers)
    for row in ROWS:
        ws.append(
            [
                row["id"],
                row["category"],
                row["question"],
                row["url"],
                row["answer"],
                row["references"],
                row["expected"],
                row["finding"],
                row["status"],
                row["priority"],
            ]
        )
    style_sheet(ws)
    widths = [12, 18, 46, 58, 36, 52, 52, 52, 18, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    return wb


def main() -> int:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
