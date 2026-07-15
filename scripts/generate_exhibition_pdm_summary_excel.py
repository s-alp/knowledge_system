from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(
    r"C:\Users\s-iwata\Desktop\knowledge_system\ナレッジシステムお披露目会向け\お披露目会_作成資料_20260601\pdm_csv_exhibition_2026-06-03\20260603_ナレッジシステム用CSV"
)
OUTPUT_PATH = BASE_DIR.parent / "展示会向けPDMデータ概要_2026-06-03.xlsx"


def resolve_csv(prefix: str) -> Path:
    matches = sorted(BASE_DIR.glob(f"{prefix}_*.csv"))
    if not matches:
        raise FileNotFoundError(f"CSV not found for prefix: {prefix}")
    return matches[0]


def read_csv(prefix: str) -> list[dict[str, str]]:
    path = resolve_csv(prefix)
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def autosize(ws, min_width: int = 10, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = []
        for cell in column_cells:
            if cell.value is None:
                continue
            values.extend(str(cell.value).splitlines())
        longest = max((len(v) for v in values), default=0)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_width, longest + 2)
        )


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def style_body(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_table(ws, table_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_sheet(ws, rows: list[dict[str, str]], title_fill: str | None = None) -> None:
    if not rows:
        ws["A1"] = "データなし"
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    style_header(ws[1])
    if title_fill:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=title_fill)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    style_body(ws)
    add_table(ws, f"tbl_{ws.title}".replace(" ", "_").replace("-", "_"))
    autosize(ws)


def build_overview_rows(
    projects, products, parts, drawings, documents, project_products, product_parts
) -> list[list[str | int]]:
    shared_products = [p for p in products if p["ライン名"] == "共通モジュール"]
    shared_product_counts = Counter(
        link["製品・装置・ユニット名"] for link in project_products
    )
    shared_part_counts = Counter(link["部品番号"] for link in product_parts)
    purchased_parts = [p for p in parts if p["仕入先"] and p["仕入先"] != "内製"]
    inhouse_parts = [p for p in parts if p["仕入先"] == "内製"]

    return [
        ["項目", "値", "補足"],
        ["プロジェクト数", len(projects), "展示会向け創作案件"],
        ["製品・装置・ユニット数", len(products), "共通モジュールを含む"],
        ["共通製品数", len(shared_products), "複数案件で流用する標準モジュール"],
        ["部品数", len(parts), "購入品と内製品を混在"],
        ["購入品数", len(purchased_parts), "仕入先が入っている部品"],
        ["内製部品数", len(inhouse_parts), "仕入先が `内製` の部品"],
        ["図面数", len(drawings), "実ファイル連携不要前提の創作図面"],
        ["文書数", len(documents), "仕様書・設計書・マニュアル"],
        [
            "最も共有される製品",
            shared_product_counts.most_common(1)[0][0],
            f"{shared_product_counts.most_common(1)[0][1]}案件で利用",
        ],
        [
            "最も共有される部品",
            shared_part_counts.most_common(1)[0][0],
            f"{shared_part_counts.most_common(1)[0][1]}製品で利用",
        ],
        [
            "公開可否",
            "問題ない想定",
            "会社名・人名・案件名・製品名は創作。実ファイルパスは空欄。",
        ],
    ]


def rows_to_dicts(matrix: list[list[str | int]]) -> list[dict[str, str]]:
    headers = [str(x) for x in matrix[0]]
    rows = []
    for body in matrix[1:]:
        rows.append({headers[i]: str(body[i]) for i in range(len(headers))})
    return rows


def build_shared_product_rows(project_products, products) -> list[dict[str, str]]:
    counts = Counter(link["製品・装置・ユニット名"] for link in project_products)
    product_map = {p["製品・装置・ユニット名"]: p for p in products}
    rows = []
    for name, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        product = product_map[name]
        rows.append(
            {
                "製品・装置・ユニット名": name,
                "カテゴリ": product["カテゴリ"],
                "種別": product["種別"],
                "ライン名": product["ライン名"],
                "利用プロジェクト数": str(count),
                "共有区分": "共通モジュール" if product["ライン名"] == "共通モジュール" else "案件固有",
            }
        )
    return rows


def build_shared_part_rows(product_parts, parts) -> list[dict[str, str]]:
    counts = Counter(link["部品番号"] for link in product_parts)
    part_map = {p["部品番号"]: p for p in parts}
    rows = []
    for part_no, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        part = part_map[part_no]
        rows.append(
            {
                "部品番号": part_no,
                "部品名": part["部品名"],
                "カテゴリ": part["カテゴリ"],
                "仕入先": part["仕入先"],
                "共有製品数": str(count),
                "区分": "購入品" if part["仕入先"] and part["仕入先"] != "内製" else "内製品",
            }
        )
    return rows


def build_project_view_rows(projects, project_products) -> list[dict[str, str]]:
    grouped = defaultdict(list)
    for link in project_products:
        grouped[link["プロジェクト名"]].append(link["製品・装置・ユニット名"])

    rows = []
    for project in projects:
        product_names = grouped[project["プロジェクト名"]]
        rows.append(
            {
                "プロジェクト名": project["プロジェクト名"],
                "顧客名": project["顧客名"],
                "顧客担当者": project["顧客担当者"],
                "責任者": project["責任者"],
                "ステータス": project["ステータス"],
                "開始日": project["開始日"],
                "終了予定日": project["終了予定日"],
                "主な製品構成": " / ".join(product_names[:6]),
            }
        )
    return rows


def build_story_rows() -> list[dict[str, str]]:
    return [
        {
            "順番": "1",
            "観点": "世界観",
            "説明": "蒼雲機装株式会社という設備メーカーが、複数の製造業顧客向けに専用設備と共通モジュールを供給している前提のデータセットです。",
        },
        {
            "順番": "2",
            "観点": "顧客構成",
            "説明": "顧客は企業単位ではなく、実務上やり取りする部署単位で登録しています。同一企業の別部署が別案件を持つ構成です。",
        },
        {
            "順番": "3",
            "観点": "ユーザー",
            "説明": "社内担当者は実画面の考え方に合わせてユーザーマスタとして整理し、姓名分割、メールアドレス、部署、役職を持たせています。案件や製品の責任者・担当者はこの表示名を参照します。",
        },
        {
            "順番": "4",
            "観点": "プロジェクト",
            "説明": "プロジェクトは量産設備や検査セルなど、設備メーカーが受注する案件単位で整理しています。各案件には客先担当者と社内責任者を割り当てています。",
        },
        {
            "順番": "5",
            "観点": "製品・装置・ユニット",
            "説明": "製品は案件固有の主設備と、複数案件で再利用される共通モジュールに分けています。安全囲い、制御盤、搬送コンベアなどは標準化した自社モジュールの想定です。",
        },
        {
            "順番": "6",
            "観点": "共有製品",
            "説明": "共通モジュールは複数案件へ横展開されるため、プロジェクトごとにゼロから設計しない構造になっています。共有製品シートで利用案件数を確認できます。",
        },
        {
            "順番": "7",
            "観点": "部品",
            "説明": "部品は購入品と内製品を混在させています。標準ボルトやセンサは購入品、ブラケットやフレーム、専用プレート類は内製品として扱っています。",
        },
        {
            "順番": "8",
            "観点": "共有部品",
            "説明": "標準部品は複数製品で使い回し、内製部品は主に個別製品へ紐づく構成です。共有部品シートで、どの部品が何製品で再利用されるかを見られます。",
        },
        {
            "順番": "9",
            "観点": "図面",
            "説明": "図面は製品代表図、部品図、レイアウト図を持たせています。実ファイル連携は不要前提なので、図面番号・タイトル・用途が分かる粒度にしています。",
        },
        {
            "順番": "10",
            "観点": "文書",
            "説明": "文書は要件仕様書、レイアウト検討書、運転手順書、保守点検手順書、構想設計書など、設備メーカーが実務で持つ文書種別に寄せています。",
        },
        {
            "順番": "11",
            "観点": "共有状況",
            "説明": "共有状況は、共通モジュールや標準部品がどの案件・製品へ横展開されているかを見るためのビューです。展示会では標準化や再利用の説明に使えます。",
        },
        {
            "順番": "12",
            "観点": "紐づけ一覧",
            "説明": "最後の各紐づけシートを見ると、案件から製品、製品から部品、さらに図面・文書まで、どこで共有しどこが専用かを追える構成です。",
        },
    ]


def main() -> None:
    contacts = read_csv("02")
    users = read_csv("03")
    projects = read_csv("05")
    products = read_csv("06")
    parts = read_csv("07")
    drawings = read_csv("08")
    documents = read_csv("09")
    project_products = read_csv("10")
    product_hierarchy = read_csv("11")
    product_parts = read_csv("12")
    project_drawings = read_csv("13")
    product_drawings = read_csv("14")
    part_drawings = read_csv("15")
    project_documents = read_csv("16")
    product_documents = read_csv("17")
    part_documents = read_csv("18")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    overview_ws = wb.create_sheet("概要")
    write_sheet(
        overview_ws,
        rows_to_dicts(
            build_overview_rows(
                projects,
                products,
                parts,
                drawings,
                documents,
                project_products,
                product_parts,
            )
        ),
        title_fill="0B6E4F",
    )

    sheets = [
        ("ストーリー", build_story_rows(), "375623"),
        ("顧客担当者一覧", contacts, "2F5597"),
        ("ユーザー一覧", users, "548235"),
        ("プロジェクト一覧", build_project_view_rows(projects, project_products), "1F4E78"),
        ("製品一覧", products, "7F6000"),
        ("部品一覧", parts, "5B9BD5"),
        ("図面一覧", drawings, "7030A0"),
        ("文書一覧", documents, "C55A11"),
        ("共有製品", build_shared_product_rows(project_products, products), "2F75B5"),
        ("共有部品", build_shared_part_rows(product_parts, parts), "4F81BD"),
        ("PJ_製品", project_products, "3C78D8"),
        ("製品_親子", product_hierarchy, "8064A2"),
        ("製品_部品", product_parts, "70AD47"),
        ("PJ_図面", project_drawings, "9E480E"),
        ("製品_図面", product_drawings, "A64D79"),
        ("部品_図面", part_drawings, "5B5EA6"),
        ("PJ_文書", project_documents, "2E8B57"),
        ("製品_文書", product_documents, "2E8B57"),
        ("部品_文書", part_documents, "2E8B57"),
    ]

    for title, rows, color in sheets:
        ws = wb.create_sheet(title)
        write_sheet(ws, rows, color)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    wb.save(OUTPUT_PATH)
    print(str(OUTPUT_PATH))


if __name__ == "__main__":
    main()
