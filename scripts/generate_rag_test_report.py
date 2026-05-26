from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
OUTPUT_DIR = WORKSPACE / "output" / "spreadsheet"
OUTPUT_PATH = OUTPUT_DIR / "rag_test_summary_reconciled_2026-04-23.xlsx"


SUMMARY_ROWS = [
    ("検証日", "2026-04-23 09:43:47"),
    ("対象URL", "http://210.165.3.139/web/chat"),
    ("検証方式", "ブラックボックス検証（チャットUI経由）"),
    ("主な対象", "コマツ小山の治具 / コマツ小山のガントリー / 広島アルミ / 澁谷工業規格"),
    ("主要な問題傾向", "略語解釈の失敗、案件混線、負例時の誤根拠、BOM系での案件粒度不足"),
]


TEST_ROWS = [
    {
        "id": "T01",
        "category": "規格検索",
        "question": "SESの熱処理に関する規格を教えてください。根拠になった資料名も示してください。",
        "chat_url": "http://210.165.3.139/web/chat/0ed8ad53-6c09-4b7c-833c-20dc0a36e284",
        "result": "SESを「表面エンジニアリングサービス」と誤解しつつ、熱処理系資料を一部参照。",
        "fact": "略語SESを社内規格名として解釈できず、回答本文が一般語義に流れた。根拠には熱処理資料が含まれる一方で、質問意図と回答本文が不整合だった。",
        "proposal": "SESと管理番号資料を結ぶ別名辞書を導入する。回答生成前に『質問中の略語』と『参照資料カテゴリ』の整合チェックを入れる。",
        "dev_question": "略語辞書や別名辞書をRAG検索前に適用できますか。SESのような社内規格名を正式資料へ寄せる仕組みはありますか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T02",
        "category": "規格検索",
        "question": "澁谷工業のSES規格で、M5に対するボルト穴サイズを教えてください。根拠資料名とページも示してください。",
        "chat_url": "http://210.165.3.139/web/chat/084274c8-31f5-453b-a40f-b4f71bc67688",
        "result": "6 mmと回答したが、澁谷工業資料に加えてコマツ小山の治具チェックリストが根拠に混在。",
        "fact": "数値自体は妥当でも、根拠に別案件の資料が混ざった。『澁谷工業』と明示しても案件境界の切り分けに失敗している。",
        "proposal": "案件名・企業名のメタデータフィルタを再ランキング前に強制適用する。回答に使う根拠集合から質問条件と一致しない資料を除外する。",
        "dev_question": "企業名・案件名の条件を満たさない資料を根拠候補から除外する設定は可能ですか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T03",
        "category": "案件検索",
        "question": "コマツ小山のガントリーの走行速度を教えてください。根拠資料名とページも示してください。",
        "chat_url": "http://210.165.3.139/web/chat/7fa9e43e-db36-4c88-8b55-082061194028",
        "result": "60 m/min (= 1 m/s) と回答し、コマツ小山ガントリー資料の計算シートを参照。",
        "fact": "案件名が明確で、資料パスにも案件名・ガントリーが含まれる場合は比較的安定して正答できた。",
        "proposal": "このケースの検索条件を成功パターンとして分析し、案件名とカテゴリ名の一致を他クエリにも転用する。",
        "dev_question": "",
        "severity": "低",
        "status": "良好",
    },
    {
        "id": "T04",
        "category": "負例判定",
        "question": "トヨタ向けのガントリー資料はありますか。根拠資料名も示してください。見つからない場合は、ないと明言してください。",
        "chat_url": "http://210.165.3.139/web/chat/99c050f4-c267-4438-b03d-99fc30740dd0",
        "result": "『ない』と回答したが、参考にコマツ小山の治具チェックリストを表示。",
        "fact": "負例判定そのものはできても、根拠が無関係資料だった。『ない』と答える際の根拠選定ロジックが破綻している。",
        "proposal": "負例時は『根拠なし』を許容し、無関係資料を参考に出さない。質問条件と参考資料が一致しない場合は参考欄を空にするガードを入れる。",
        "dev_question": "資料無しのときに『参考なし』で返す設計は可能ですか。それとも現状は必ず何か参考を出す仕様ですか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T05",
        "category": "案件検索",
        "question": "広島アルミの治具資料はありますか。見つからない場合は、ないと明言し、参考資料名だけ示してください。",
        "chat_url": "http://210.165.3.139/web/chat/325bf083-bbc7-4a38-92ea-112963963ff6",
        "result": "『ある』と回答し、広島アルミ配下のチェックリストを根拠として提示。",
        "fact": "客先名レベルでは広島アルミに着地できている。ただし、回答本文の『治具図と治具カバー図面が展開』という記述は、チェックリストだけからどこまで言えるかが曖昧だった。",
        "proposal": "回答本文で断定する情報は、根拠資料の記載範囲に限定する。チェックリスト由来の情報と図面そのものの情報を区別して表現する。",
        "dev_question": "チェックリスト由来の情報だけで図面展開や設計内容まで断定しないよう、回答テンプレートを制御できますか。",
        "severity": "中",
        "status": "要改善",
    },
    {
        "id": "T06",
        "category": "規格検索",
        "question": "澁谷工業のSES規格の資料名を列挙してください。コマツ資料は含めないでください。",
        "chat_url": "http://210.165.3.139/web/chat/ceffbcf7-3768-4204-bd70-6d7526ee97ae",
        "result": "『列挙できない』と回答しつつ、澁谷工業の電気用図記号PDFを参考表示。",
        "fact": "企業名までは合っているが、規格カテゴリが外れて電気資料に流れた。『コマツ資料を含めない』条件は守れても、『SES規格』条件は守れていない。",
        "proposal": "企業フィルタだけでなく、質問語と資料タイトル・フォルダカテゴリの意味一致スコアを強化する。カテゴリ不一致の資料は参考欄から除外する。",
        "dev_question": "同一企業内でもカテゴリ不一致資料を参考から落とす設定はできますか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T07",
        "category": "資料名直接指定",
        "question": "澁谷工業の「熱処理指定方法_2GDE82D.pdf」はありますか。あれば資料名とページだけ答えてください。",
        "chat_url": "http://210.165.3.139/web/chat/a1826c7b-98d6-4b17-99a2-fc2f18279634",
        "result": "『資料は存在します』と回答し、資料名とページ1を正しく返答。",
        "fact": "正式資料名を直接指定した場合は正しくヒットした。別名・略語よりも、ファイル名ベースの検索は強い。",
        "proposal": "正式資料名が分からない利用者向けに、別名・略語から正式資料名へ寄せる辞書・候補提示を用意する。",
        "dev_question": "",
        "severity": "低",
        "status": "良好",
    },
    {
        "id": "T08",
        "category": "BOM/部品横断",
        "question": "SMCのシリンダーが使われている案件を挙げてください。根拠資料名も示してください。",
        "chat_url": "http://210.165.3.139/web/chat/e1d985e6-c078-41b5-83ae-600020158621",
        "result": "『カバー一般のエアシリンダー（SMC製CG1フート型タイプ）』と回答し、広島アルミのチェックリストを根拠表示。",
        "fact": "質問は案件名列挙だったが、返答は用途・部品名に留まり、案件粒度へ変換できなかった。BOM情報から案件情報へ上位概念を引く処理が弱い。",
        "proposal": "部品名→構成→案件名のメタデータ連結を強化する。質問が『案件を挙げて』のときは、案件名が取れない場合に『案件名は抽出できない』と返す制御を入れる。",
        "dev_question": "BOM系検索で『部品名は取れるが案件名へ上げられない』場合、案件名未特定として返す制御は可能ですか。",
        "severity": "中",
        "status": "要改善",
    },
    {
        "id": "T09",
        "category": "BOM/部品横断",
        "question": "案件名を答えてください。部品名や用途名ではなく、客先名や案件名を挙げてください。SMCのシリンダーが使われている案件だけを列挙してください。",
        "chat_url": "http://210.165.3.139/web/chat/e1d985e6-c078-41b5-83ae-600020158621",
        "result": "広島アルミ / コマツ小山 / No.92 を列挙したが、根拠は広島アルミの同一チェックリスト2件のみ。",
        "fact": "追撃で案件名らしき列挙はできたが、提示根拠は広島アルミ資料のみで、コマツ小山やNo.92を裏付けていない。案件抽出結果と根拠集合の整合が崩れている。",
        "proposal": "案件列挙時は、各案件に少なくとも1つずつ対応する根拠を必須にする。根拠が取れない案件は列挙対象から除外する。",
        "dev_question": "列挙した案件ごとに根拠を1件以上必須にするガードは入れられますか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T10",
        "category": "負例判定",
        "question": "澁谷工業のガントリー資料はありますか。見つからない場合は、ないと明言してください。参考は澁谷工業資料だけにしてください。",
        "chat_url": "http://210.165.3.139/web/chat/0ebb05b8-07b4-4562-9444-329bc58e23df",
        "result": "『確認できませんでした』と回答したが、参考に図番規定PDFを提示。",
        "fact": "企業フィルタは守れている一方、ガントリー無関係の『図番規定_2GAA02D.pdf』を参考に出した。負例時に無関係な同一企業資料へ逃げる傾向が残っている。",
        "proposal": "負例時は『同一企業なら何でも参考に出す』挙動を止める。質問トピックとタイトル・カテゴリが一致しない資料は参考欄から除外する。",
        "dev_question": "『資料なし』時に同一企業の別資料を参考として付けてしまうのは仕様ですか。止められますか。",
        "severity": "中",
        "status": "要改善",
    },
    {
        "id": "T11",
        "category": "案件検索",
        "question": "広島アルミのガントリー関連資料名を列挙してください。資料名だけを挙げてください。",
        "chat_url": "http://210.165.3.139/web/chat/e143eea3-2152-434d-9af5-6e03abca5fd7",
        "result": "『干渉チェックエビデンスシート_ソクブカバークミ.xlsx』を1件だけ列挙。",
        "fact": "広島アルミ配下には着地しているが、『ガントリー関連』という条件に対して側部カバーの干渉チェック資料を返しており、カテゴリ判定が甘い。",
        "proposal": "質問語『ガントリー』と資料タイトル・フォルダ名の意味一致を強化する。『カバー』など別カテゴリ資料は再ランキングで減点する。",
        "dev_question": "フォルダ名やファイル名ベースで『ガントリー』『カバー』のカテゴリ重み付けは調整できますか。",
        "severity": "中",
        "status": "要改善",
    },
    {
        "id": "T12",
        "category": "BOM/部品横断",
        "question": "KOSMEKの油圧シリンダを使っている案件を挙げてください。案件名だけではなく、根拠資料名も示してください。",
        "chat_url": "http://210.165.3.139/web/chat/dc32be1e-d259-427c-8277-1a55be832e39",
        "result": "コマツ小山（客先）に加えて、20241205_BRG組付け不具合を案件名として列挙。",
        "fact": "BOM系検索で案件と不具合件名が混在した。根拠もコマツ小山ガントリー資料と広島アルミの確認表が混ざっており、案件名粒度が不統一だった。",
        "proposal": "案件名・客先名・不具合件名・作業フォルダ名を別メタデータとして区別する。『案件を挙げて』のときは案件テーブルに正規化された名称だけを返す。",
        "dev_question": "案件名・客先名・不具合件名を別フィールドとして扱う設計にできますか。",
        "severity": "高",
        "status": "問題あり",
    },
    {
        "id": "T13",
        "category": "追撃改善",
        "question": "ここでいうSESは澁谷工業の社内規格のことです。SESを一般用語として解釈しないでください。澁谷工業の熱処理規格だけに限定して、資料名とページを答えてください。",
        "chat_url": "http://210.165.3.139/web/chat/0ed8ad53-6c09-4b7c-833c-20dc0a36e284",
        "result": "『熱処理指定方法_2GDE82D.pdf、ページ1』と回答。初回より大幅改善したが、参考にカムの材料と熱処理_2GDC82D.pdf も混在。",
        "fact": "追撃で略語誤解は解消し、正式資料へ到達できた。一方で『澁谷工業の熱処理規格だけ』という制約に対し、追加の近傍資料がまだ混ざった。",
        "proposal": "追撃で条件が明示された場合は、直前回答よりも新しい制約を強く反映する。追撃後は参考資料集合を再計算し、条件不一致資料を落とす。",
        "dev_question": "追撃で条件を追加した場合、初回の根拠集合を引きずらずに再検索・再絞り込みする仕様にできますか。",
        "severity": "中",
        "status": "改善あり",
    },
]


def apply_header_style(ws, row_idx: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_body_style(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_workbook() -> Workbook:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "サマリ"
    ws_summary.append(["項目", "内容"])
    for key, value in SUMMARY_ROWS:
        ws_summary.append([key, value])
    apply_header_style(ws_summary, 1)
    apply_body_style(ws_summary)
    set_widths(ws_summary, {"A": 24, "B": 120})

    ws_tests = wb.create_sheet("検証結果")
    headers = [
        "テストID",
        "カテゴリ",
        "質問",
        "チャットURL",
        "回答要約",
        "事実",
        "改善提案",
        "開発先確認事項",
        "重要度",
        "判定",
    ]
    ws_tests.append(headers)
    for row in TEST_ROWS:
        ws_tests.append(
            [
                row["id"],
                row["category"],
                row["question"],
                row["chat_url"],
                row["result"],
                row["fact"],
                row["proposal"],
                row["dev_question"],
                row["severity"],
                row["status"],
            ]
        )
    apply_header_style(ws_tests, 1)
    apply_body_style(ws_tests)
    set_widths(
        ws_tests,
        {
            "A": 10,
            "B": 16,
            "C": 42,
            "D": 78,
            "E": 42,
            "F": 54,
            "G": 56,
            "H": 52,
            "I": 10,
            "J": 10,
        },
    )

    for row_idx in range(2, ws_tests.max_row + 1):
        url_cell = ws_tests[f"D{row_idx}"]
        url_cell.hyperlink = url_cell.value
        url_cell.style = "Hyperlink"

    ws_tests.freeze_panes = "A2"
    ws_summary.freeze_panes = "A2"

    ws_actions = wb.create_sheet("優先対応")
    ws_actions.append(["優先度", "改善テーマ", "狙い", "補足"])
    action_rows = [
        ("P1", "別名辞書の導入", "SESなどの略語から正式資料へ寄せる", "SES -> 2GDE82D / 2GDE81D のような別名管理"),
        ("P1", "案件・企業メタデータ絞り込み", "別案件資料の混線を防ぐ", "澁谷工業質問でコマツ資料が混ざらないようにする"),
        ("P1", "負例時の参考制御", "『ない』のときに無関係根拠を出さない", "条件不一致資料は参考欄を空にする"),
        ("P2", "回答前整合チェック", "質問条件と参考資料の一致を確認する", "企業名・案件名・カテゴリ名の一致チェック"),
        ("P2", "BOMから案件名への連結強化", "部品ベース質問で案件名を返せるようにする", "部品名 -> 構成 -> 案件名 の関連付け"),
    ]
    for action in action_rows:
        ws_actions.append(action)
    apply_header_style(ws_actions, 1)
    apply_body_style(ws_actions)
    set_widths(ws_actions, {"A": 10, "B": 28, "C": 30, "D": 54})

    ws_pending = wb.create_sheet("未回収整理")
    ws_pending.append(["種別", "内容", "状態", "補足"])
    pending_rows = [
        (
            "未回収",
            "広島アルミのガントリー関連資料として、側部カバーや干渉チェックを除外する追撃",
            "未回収",
            "送信済みだが回答本文をまだ回収していない",
        ),
        (
            "未回収",
            "SMC案件に対する追加の出力形式指定追撃",
            "未回収",
            "追撃比較の途中。案件名と根拠の1対1対応を確認したい",
        ),
        (
            "整理対象外",
            "http://210.165.3.139/web/chat の素の重複タブ群",
            "除外",
            "質問未送信のタブが混ざるため、今回の回収済み版では管理対象から外す",
        ),
    ]
    for row in pending_rows:
        ws_pending.append(row)
    apply_header_style(ws_pending, 1)
    apply_body_style(ws_pending)
    set_widths(ws_pending, {"A": 14, "B": 54, "C": 12, "D": 52})

    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 36

    return wb


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
