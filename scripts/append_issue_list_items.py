from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
TARGET_PATH = WORKSPACE / "ナレッジシステム_やり取り項目リスト.xlsx"
SHEET_NAME = "検索関連やり取り項目リスト"

NEW_ITEMS = [
    (
        31,
        "追撃条件を足しても、追加条件に合う資料へ十分に再絞り込みされない",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/e143eea3-2152-434d-9af5-6e03abca5fd7\n\n"
        "確認内容:\n"
        "『広島アルミのガントリー関連資料として、側部カバーや干渉チェックは除外してください。ガントリーまたはハンド関係の資料名だけを列挙してください。』と追撃したが、"
        "回答は広島アルミ側のガントリー資料ではなく、コマツ小山側の『20220711_ガントリーハンド資料/新しいテキスト ドキュメント.txt』へ飛んだ。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\hiroshima_alumi_gantry\\現地改造計画図.pdf\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\hiroshima_alumi_gantry\\ケースライン確認図.pdf\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/治具/第二期治具/受領/20220711_ガントリーハンド資料/新しいテキスト ドキュメント.txt\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/ガントリー/受領/受領資料/FIX段替考え方.xlsx\n\n"
        "除外条件や客先条件を追加しても、初回の検索・根拠集合を十分に捨て切れず、追加条件に合う資料へ再絞り込みされていないように見える。",
        "関連No: 25, 30",
    ),
    (
        32,
        "案件名だけを求めた際に、列挙した案件と根拠資料の対応が取れていない",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/e1d985e6-c078-41b5-83ae-600020158621\n\n"
        "確認内容:\n"
        "『案件名を答えてください。部品名や用途名ではなく、客先名や案件名を挙げてください。SMCのシリンダーが使われている案件だけを列挙してください。』と指示したところ、"
        "広島アルミ・コマツ小山・No.92 と列挙されたが、根拠資料は広島アルミ側部カバーのチェックリスト 2 件のみだった。\n\n"
        "正解ファイル:\n"
        "案件名を直接裏付けられる BOM / 構成表 / チェックリストが案件ごとに最低1件ずつ必要。\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/広島アルミ/20241101_カバー関係、ロボット関係/作業/01_側部カバー/C500-12_カバークミ品質要求事項確認チェックリスト.xlsx\n"
        "（同一資料の別シートのみで、コマツ小山や No.92 は裏付けられていない）\n\n"
        "案件名の列挙と根拠の対応関係が崩れているため、案件名を返すなら各案件を直接裏付ける資料を提示してほしい。",
        "関連No: 18, 25",
    ),
    (
        33,
        "規格表の列挙項目を正しく抜かず、表とずれた要約になる",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/568f633a-3261-48f5-945f-e6299d40498d\n\n"
        "確認内容:\n"
        "『澁谷工業の熱処理指定方法で、部品図に指定すべき項目を列挙してください。資料名とページも示してください。』と質問したところ、"
        "回答は『熱処理箇所の明示』『指定名称の記載』となった。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\shibuya_specs\\熱処理指定方法_2GDE82D.pdf\n"
        "（表1『指示すべき項目 と 指定方法』の列挙対象は、熱処理・硬度・硬化深さ）\n\n"
        "不正解ファイル:\n"
        "正しい資料には着地しているが、表の列挙項目をそのまま抜かず、備考相当の内容へ言い換わっている。\n\n"
        "資料そのものは合っているが、表の列挙項目を正しく抽出できていない。",
        "関連No: 27",
    ),
    (
        34,
        "改造対応項目の列挙で、本来の対応項目メモではなく別チェックリストへ誤着地する",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/fb103f1e-cd92-4132-9404-a8344bb75323\n\n"
        "確認内容:\n"
        "『広島アルミの改造対応項目として挙がっているものを列挙してください。根拠資料名も示してください。』と質問したところ、"
        "回答は『戻しボタン』『切削条件』『製図作業』で、根拠はツーリング品質要求事項確認チェックリストだった。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\hiroshima_alumi_gantry\\20241202_対応項目.txt\n"
        "（ワークハンド、ハンド置台、踏台、安全柵、ロボット制御盤架台、トイ、クーラント2次タンク移設用トイ・配管）\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/広島アルミ/20241101_カバー関係、ロボット関係/受領/20250107_ツーリング資料と見積照会/C300-09_ツーリング品質要求事項確認チェックリスト.xlsx\n\n"
        "『対応項目』の列挙なのに、質問に最も近いテキストメモではなく別用途のチェックリストへ飛んでいる。",
        "関連No: 25",
    ),
    (
        35,
        "Excel仕様書の対象シート・セルではなく別資料へ誤着地し、回答が逆転する",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/45cc17d9-c03f-442b-8185-168423b38940\n\n"
        "確認内容:\n"
        "『コマツ小山のガントリーハンド仕様で、車種107の前半はSTEP3移行時にツメの新製が必要ですか。根拠資料名も示してください。』と質問したところ、"
        "回答は『不要』だった。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\excel_word_specs\\komatsu_koyama\\STEP2_＞STEP3ハンドの仕様r2.xlsx\n"
        "（車種107 前半の『STEP3移行時 ツメの新製』は『必要』）\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/ガントリー/資料/220401_コマツ小山__スマートライン2_確認・決定必要項目.xlsx\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/基礎図/Step1_3比較_O2案_210803.pdf\n\n"
        "質問に最も近い Excel 仕様表の対象セルへ行かず、別資料を横断して逆の回答になっている。",
        "関連No: 13, 25",
    ),
    (
        36,
        "Excel仕様書の値は合っても、参考資料が過剰に混線する",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/11de51e7-c775-46b2-a834-ba98b918821c\n\n"
        "確認内容:\n"
        "『広島アルミのミスト・集塵装置クミのPJ No.、形格、組図名称を教えてください。根拠資料名も示してください。』と質問したところ、"
        "PJ No.=00K1A4980、形格=FTL-2184GC、組図名称=ミスト・集塵装置クミ と回答し、値自体は合っていた。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\excel_word_specs\\hiroshima_alumi\\C800-02_ミスト・集塵装置クミ品質要求事項確認チェックリスト.xlsx\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/広島アルミ/20240508_N4H6_マガジン改造/受領/20240726_特殊マガジンについて確認依頼など/20240726_依頼内容/確認表.xlsx\n"
        "/mnt/crawl/アルパイン/広島アルミ/20240508_N4H6_マガジン改造/受領/20241205_BRG組付け不具合/確認表.xlsx\n"
        "/mnt/crawl/アルパイン/広島アルミ/20241101_カバー関係、ロボット関係/作業/07_ツーリング/20250129_OP40L加工確認資料.xlsx\n\n"
        "答えは合っていても、質問に直接対応する仕様書だけで足りる内容に対して、別案件や別カテゴリの資料まで参考に混ぜている。",
        "関連No: 13, 25",
    ),
    (
        37,
        "Word仕様書の本文数値を正しく拾えず、別資料へ誤着地する",
        "2026-04-23",
        "URL:\n"
        "http://210.165.3.139/web/chat/6357f591-8477-4b47-8b20-bf25f7c77893\n\n"
        "確認内容:\n"
        "『コマツ小山ガントリーの保守要領で、X軸走行用サーボモーター交換時のバックラッシュ狙い値はいくつですか。根拠資料名も示してください。』と質問したところ、"
        "回答は『0.2』だった。\n\n"
        "正解ファイル:\n"
        "C:\\Users\\s-iwata\\Desktop\\knowledge_system\\local_test_materials\\excel_word_specs\\komatsu_koyama\\maintenance_（gantry）-スマートライン2_最新.doc\n"
        "（本文に『バックラッシュ0.45mm狙い』と記載あり）\n\n"
        "不正解ファイル:\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/ガントリー/資料/220401_コマツ小山__スマートライン2_確認・決定必要項目.xlsx\n"
        "/mnt/crawl/アルパイン/コマツ/コマツ小山/ガントリー/受領/20220803_シリンダー諸元作成用資料/KO小山ガントリシリンダー・潤滑一覧.xlsx\n\n"
        "Word 手順書の見出し・段落本文よりも、別の Excel 資料が優先されているように見えるため、Word 文書本文の数値と単位を優先して取得できるか確認したい。",
        "関連No: 25, 30",
    ),
]

ROW_MAP = {
    31: 182,
    32: 184,
    33: 186,
    34: 188,
    35: 190,
    36: 192,
    37: 194,
}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def populate_item_row(ws, row: int, item_data: tuple[int, str, str, str, str]) -> None:
    no, item, confirm_date, content, remarks = item_data
    ws.cell(row, 1).value = no
    ws.cell(row, 2).value = item
    ws.cell(row, 3).value = "起票"
    ws.cell(row, 4).value = "アルパイン"
    ws.cell(row, 5).value = confirm_date
    ws.cell(row, 6).value = content
    ws.cell(row, 7).value = remarks


def clear_row(ws, row: int) -> None:
    for col in range(1, ws.max_column + 1):
        try:
            ws.cell(row, col).value = None
        except AttributeError:
            pass


def main() -> int:
    wb = load_workbook(TARGET_PATH)
    ws = wb[SHEET_NAME]

    template_item_row = 176
    template_blank_row = 177

    for row in range(182, 196):
        clear_row(ws, row)

    for no, item, confirm_date, content, remarks in NEW_ITEMS:
        row = ROW_MAP[no]
        copy_row_style(ws, template_item_row, row)
        copy_row_style(ws, template_blank_row, row + 1)
        populate_item_row(ws, row, (no, item, confirm_date, content, remarks))
        clear_row(ws, row + 1)

    wb.save(TARGET_PATH)
    print(TARGET_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
