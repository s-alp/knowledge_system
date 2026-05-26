from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
import fitz


WORKSPACE = Path(r"C:\Users\s-iwata\Desktop\knowledge_system")
ROOT = WORKSPACE / "local_test_materials"


def read_pdf(path: Path) -> dict[str, object]:
    doc = fitz.open(path)
    pages: list[dict[str, object]] = []
    for page_index in range(min(3, doc.page_count)):
        page = doc.load_page(page_index)
        text = page.get_text() or ""
        pages.append(
            {
                "page": page_index + 1,
                "text_excerpt": " ".join(text.split())[:2000],
            }
        )
    return {
        "type": "pdf",
        "page_count": doc.page_count,
        "pages": pages,
    }


def read_xlsx(path: Path) -> dict[str, object]:
    wb = load_workbook(path, read_only=True, data_only=False)
    sheets: dict[str, object] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            rows.append([None if value is None else str(value) for value in row])
        sheets[sheet_name] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sample_rows": rows,
        }
    return {
        "type": "xlsx",
        "sheets": sheets,
    }


def read_txt(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8")
    return {
        "type": "txt",
        "text": text,
    }


def main() -> int:
    result: dict[str, object] = {}
    for path in sorted(ROOT.rglob("*")):
        if not path.is_file():
            continue
        rel = path.relative_to(WORKSPACE).as_posix()
        if path.suffix.lower() == ".pdf":
            result[rel] = read_pdf(path)
        elif path.suffix.lower() == ".xlsx":
            result[rel] = read_xlsx(path)
        elif path.suffix.lower() == ".txt":
            result[rel] = read_txt(path)

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
